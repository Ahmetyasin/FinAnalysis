from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Header
from fastapi.responses import JSONResponse
import openpyxl
import json
import numpy as np
import tempfile


app = FastAPI()


def valid_api_token(api_token: str = Header(None)):
    expected_token = "2BC56DDH"  # This should be a secure, secret token
    if api_token != expected_token:
        raise HTTPException(status_code=400, detail="Invalid API Token")


# Define the two sets of values
keys = [
    "Dönen Varlıklar",
    "Nakit ve Nakit Benzerleri",
    "Finansal Yatırımlar",
    "Ticari Alacaklar",
    "Finans Sektörü Faaliyetlerinden Alacaklar",
    "Diğer Alacaklar",
    "Stoklar",
    "Peşin Ödenmiş Giderler",
    "Diğer Dönen Varlıklar",
    "Satış Amacıyla Elde Tutulan Duran Varlıklar",
    "(Ara Toplam) 1",
    "Duran Varlıklar",
    "Ticari Alacaklar",
    "Finans Sektörü Faaliyetlerinden Alacaklar",
    "Finansal Yatırımlar",
    "Diğer Alacaklar",
    "Maddi Duran Varlıklar",
    "Maddi Olmayan Duran Varlıklar",
    "Peşin Ödenmiş Giderler",
    "Ertelenmiş Vergi Varlığı",
    "Diğer Duran Varlıklar",
    "Kullanım Hakkı Varlıkları",
    "Yatırım Amaçlı Gayrimenkuller",
    "Özkaynak Yöntemiyle Değerlenen Yatırımlar",
    "TOPLAM VARLIKLAR",
    "Kısa Vadeli Yükümlülükler",
    "Finansal Borçlar",
    "Diğer Finansal Yükümlülükler",
    "Ticari Borçlar",
    "Finans Sektörü Faaliyetlerinden Borçlar",
    "Çalışanlara Sağlanan Faydalar Kapsamında Borçlar",
    "Diğer Borçlar",
    "Devlet Teşvik ve Yardımları",
    "Ertelenmiş Gelirler",
    "Dönem Karı Vergi Yükümlülüğü",
    "Borç Karşılıkları",
    "Türev Araçlar",
    "Diğer Kısa Vadeli Yükümlülükler",
    "Müşteri Sözleşmelerinden Doğan Yükümlülükler",
    "Satış Amaçlı Sınıflandırılan Varlık Gruplarına İlişkin Yükümlülükler",
    "(Ara Toplam) 2",
    "Uzun Vadeli Yükümlülükler",
    "Finansal Borçlar",
    "Ticari Borçlar",
    "Finans Sektörü Faaliyetlerinden Borçlar",
    "Diğer Borçlar",
    "Ertelenmiş Gelirler",
    "Uzun vadeli Karşılıklar",
    "Ertelenmiş Vergi Yükümlülüğü",
    "Diğer Uzun Vadeli Yükümlülükler",
    "Devlet Teşvik ve Yardımları",
    "Özkaynaklar",
    "Ana Ortaklığa Ait Özkaynaklar",
    "Ödenmiş Sermaye",
    "Kardan Ayrılan Kısıtlanmış Yedekler",
    "Geçmiş Yıllar Kar/Zararları",
    "Dönem Net Kar/Zararı",
    "Diğer Özkaynak Kalemleri",
    "Hedge Dahil Net Yabancı Para Pozisyonu",
    "Yabancı Para Çevrim Farkları",
    "Azınlık Payları",
    "TOPLAM KAYNAKLAR",
]

keys2 = [
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Dönen Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Duran Varlıklar",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Kısa Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Uzun Vadeli Yükümlülükler",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
    "Özkaynaklar",
]
# Correcting the incomplete string and completing the list of values
values = [
    "1.Dönen Varlıklar",
    "10.Nakit ve Nakit Benzerleri",
    "11.Finansal Yatırımlar",
    "12.Ticari Alacaklar",
    "12_1.Finans Sektörü Faaliyetlerinden Alacaklar",
    "13.Diğer Alacaklar",
    "15.Stoklar",
    "18.Peşin Ödenmiş Giderler",
    "19.Diğer Dönen Varlıklar",
    "19_1.Satış Amacıyla Elde Tutulan Duran Varlıklar",
    "AT1.(Ara Toplam) 1",
    "2.Duran Varlıklar",
    "22.Ticari Alacaklar",
    "22_1.Finans Sektörü Faaliyetlerinden Alacaklar",
    "24.Finansal Yatırımlar",
    "23.Diğer Alacaklar",
    "25.Maddi Duran Varlıklar",
    "26.Maddi Olmayan Duran Varlıklar",
    "28.Peşin Ödenmiş Giderler",
    "28_1.Ertelenmiş Vergi Varlığı",
    "29.Diğer Duran Varlıklar",
    "29_1.Kullanım Hakkı Varlıkları",
    "25_3.Yatırım Amaçlı Gayrimenkuller",
    "25_2.Özkaynak Yöntemiyle Değerlenen Yatırımlar",
    "TVAR.TOPLAM VARLIKLAR",
    "3.Kısa Vadeli Yükümlülükler",
    "30.Finansal Borçlar",
    "30_1.Diğer Finansal Yükümlülükler",
    "32.Ticari Borçlar",
    "32_1.Finans Sektörü Faaliyetlerinden Borçlar",
    "34.Çalışanlara Sağlanan Faydalar Kapsamında Borçlar",
    "33.Diğer Borçlar",
    "33_1.Devlet Teşvik ve Yardımları",
    "38.Ertelenmiş Gelirler",
    "36.Dönem Karı Vergi Yükümlülüğü",
    "37.Borç Karşılıkları",
    "309.Türev Araçlar",
    "390.Diğer Kısa Vadeli Yükümlülükler_",
    "309_1.Müşteri Sözleşmelerinden Doğan Yükümlülükler",
    "390_1.Satış Amaçlı Sınıflandırılan Varlık Gruplarına İlişkin Yükümlülükler",
    "AT2.(Ara Toplam) 2",
    "4.Uzun Vadeli Yükümlülükler",
    "40.Finansal Borçlar",
    "42.Ticari Borçlar",
    "40_1.Finans Sektörü Faaliyetlerinden Borçlar",
    "43.Diğer Borçlar",
    "48.Ertelenmiş Gelirler",
    "47.Uzun vadeli Karşılıklar",
    "481.Ertelenmiş Vergi Yükümlülüğü",
    "49.Diğer Uzun Vadeli Yükümlülükler",
    "49_1.Devlet Teşvik ve Yardımları",
    "5.Özkaynaklar",
    "52.Ana Ortaklığa Ait Özkaynaklar",
    "50.Ödenmiş Sermaye",
    "54.Kardan Ayrılan Kısıtlanmış Yedekler",
    "57.Geçmiş Yıllar Kar/Zararları",
    "59.Dönem Net Kar Zararı",
    "529.Diğer Özkaynak Kalemleri",
    "241.Hedge Dahil Net Yabancı Para Pozisyonu",
    "5_x.Yabancı Para Çevrim Farkları",
    "53.Azınlık Payları",
    "TKAY.TOPLAM KAYNAKLAR",
]

balance_sheet_items = list(zip(keys, keys2))

# Convert Turkish characters in values to English equivalents and replace spaces with underscores
translated_values = []
for value in values:
    translated_values.append(
        value.replace(".", "_")
        .replace("ç", "c")
        .replace("Ç", "C")
        .replace("ş", "s")
        .replace("Ş", "S")
        .replace("ğ", "g")
        .replace("Ğ", "G")
        .replace("ü", "u")
        .replace("Ü", "u")
        .replace("ö", "o")
        .replace("Ö", "O")
        .replace("ı", "i")
        .replace("İ", "I")
        .replace(" ", "_")
        .replace("(", "_")
        .replace(")", "_")
        .replace("/", "_")
    )


# Create the dictionary
balance_dict = dict(zip(balance_sheet_items, translated_values))


# List of descriptive names
descriptive_names = [
    "Satış Gelirleri",
    "Yurt İçi Satışlar",
    "Yurt Dışı Satışlar",
    "Satışların Maliyeti (-)",
    "Ticari Faaliyetlerden Brüt Kar (Zarar)",
    "BRÜT KAR (ZARAR)",
    "Genel Yönetim Giderleri (-)",
    "Pazarlama, Satış ve Dağıtım Giderleri (-)",
    "Araştırma ve Geliştirme Giderleri (-)",
    "Diğer Faaliyet Gelirleri",
    "Diğer Faaliyet Giderleri (-)",
    "FAALİYET KARI (ZARARI)",
    "Yatırım Faaliyetlerinden Gelirler",
    "Yatırım Faaliyetlerinden Giderler (-)",
    "FİNANSMAN GELİRİ (GİDERİ) ÖNCESİ FAALİYET KARI (ZARARI)",
    "(Esas Faaliyet Dışı) Finansal Gelirler",
    "(Esas Faaliyet Dışı) Finansal Giderler (-)",
    "SÜRDÜRÜLEN FAALİYETLER VERGİ ÖNCESİ KARI (ZARARI)",
    "Sürdürülen Faaliyetler Vergi Geliri (Gideri)",
    "Dönem Vergi Geliri (Gideri)",
    "Ertelenmiş Vergi Geliri (Gideri)",
    "SÜRDÜRÜLEN FAALİYETLER DÖNEM KARI/ZARARI",
    "DURDURULAN FAALİYETLER DÖNEM KARI/ZARARI",
    "DÖNEM KARI (ZARARI)",
    "Azınlık Payları",
    "Ana Ortaklık Payları",
    "Amortisman",
    "Faiz, Ücret, Prim, Komisyon ve Diğer Gelirler",
    "Faiz, Ücret, Prim, Komisyon ve Diğer Giderler (-)",
    "Finans Sektörü Faaliyetlerinden Brüt Kar (Zarar)",
    "Özkaynak Yöntemiyle Değerlenen Yatırımların Karlarından (Zararlarından) Paylar",
]

# Dictionary of keys
income_values = [
    "60_Satis_Gelirleri",
    "600_Yurt_Ici_Satislar",
    "601_Yurt_Disi_Satislar",
    "62_Satislarin_Maliyeti__e_",
    "602_Ticari_Faaliyetlerden_Brut_Kar__Zarar_",
    "BK_BRUT_KAR__ZARAR_",
    "63_Genel_Yonetim_Giderleri__e_",
    "631_Pazarlama__Satis_ve_Dagitim_Giderleri__e_",
    "630_Arastirma_ve_Gelistirme_Giderleri__e_",
    "64_Diger_Faaliyet_Gelirleri",
    "65_Diger_Faaliyet_Giderleri__e_",
    "FKZ_FAALIYET_KARI__ZARARI_",
    "64_1_Yatirim_Faaliyetlerinden_Gelirler",
    "65_1_Yatirim_Faaliyetlerinden_Giderler__e_",
    "FGOFKZ_FINANSMAN_GELIRI__GIDERI__ONCESI_FAALIYET_KARI__ZARARI_",
    "67__Esas_Faaliyet_Disi__Finansal_Gelirler",
    "68__Esas_Faaliyet_Disi__Finansal_Giderler__e_",
    "SFVOKZ_SURDURULEN_FAALIYETLER_VERGI_ONCESI_KARI__ZARARI_",
    "SFVG_Surdurulen_Faaliyetler_Vergi_Geliri__Gideri_",
    "691_Donem_Vergi_Geliri__Gideri_",
    "67_1_Ertelenmis_Vergi_Geliri__Gideri_",
    "SFDKZ_SURDURULEN_FAALIYETLER_DONEM_KARI_ZARARI",
    "DFDKZ_DURDURULAN_FAALIYETLER_DONEM_KARI_ZARARI",
    "69_DONEM_KARI__ZARARI_",
    "69_2_Azinlik_Paylari",
    "69_3_Ana_Ortaklik_Paylari",
    "796_Amortisman",
    "67_2_Faiz__Ucret__Prim__Komisyon_ve_Diger_Gelirler",
    "68_1_Faiz__Ucret__Prim__Komisyon_ve_Diger_Giderler__e_",
    "67_3_Finans_Sektoru_Faaliyetlerinden_Brut_Kar__Zarar_",
    "69_1_Ozkaynak_Yontemiyle_Deger_Yatirimlarin_Kar_Zarar__Paylar",
]


# Creating a mapping dictionary
income_dict = dict(zip(descriptive_names, income_values))


all_fields = {
    "Donem": [None, None, None, None, None, None],
    "1_Donen_Varliklar": [None, None, None, None, None, None],
    "10_Nakit_ve_Nakit_Benzerleri": [None, None, None, None, None, None],
    "12_Ticari_Alacaklar": [None, None, None, None, None, None],
    "13_Diger_Alacaklar": [None, None, None, None, None, None],
    "15_Stoklar": [None, None, None, None, None, None],
    "18_Pesin_Odenmis_Giderler": [None, None, None, None, None, None],
    "19_Diger_Donen_Varliklar": [None, None, None, None, None, None],
    "AT1__Ara_Toplam__1": [None, None, None, None, None, None],
    "2_Duran_Varliklar": [None, None, None, None, None, None],
    "24_Finansal_Yatirimlar": [None, None, None, None, None, None],
    "23_Diger_Alacaklar": [None, None, None, None, None, None],
    "25_3_Yatirim_Amacli_Gayrimenkuller": [None, None, None, None, None, None],
    "25_Maddi_Duran_Varliklar": [None, None, None, None, None, None],
    "26_Maddi_Olmayan_Duran_Varliklar": [None, None, None, None, None, None],
    "28_Pesin_Odenmis_Giderler": [None, None, None, None, None, None],
    "28_1_Ertelenmis_Vergi_Varligi": [None, None, None, None, None, None],
    "29_1_Kullanim_Hakki_Varliklari": [None, None, None, None, None, None],
    "TVAR_TOPLAM_VARLIKLAR": [None, None, None, None, None, None],
    "3_Kisa_Vadeli_Yukumlulukler": [None, None, None, None, None, None],
    "30_Finansal_Borclar": [None, None, None, None, None, None],
    "32_Ticari_Borclar": [None, None, None, None, None, None],
    "34_Calisanlara_Saglanan_Faydalar_Kapsaminda_Borclar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "33_Diger_Borclar": [None, None, None, None, None, None],
    "38_Ertelenmis_Gelirler": [None, None, None, None, None, None],
    "36_Donem_Kari_Vergi_Yukumlulugu": [None, None, None, None, None, None],
    "37_Borc_Karsiliklari": [None, None, None, None, None, None],
    "309_Turev_Araclar": [None, None, None, None, None, None],
    "309_1_Musteri_Sozlesmelerinden_Dogan_Yukumlulukler": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "390_Diger_Kisa_Vadeli_Yukumlulukler_": [None, None, None, None, None, None],
    "AT2__Ara_Toplam__2": [None, None, None, None, None, None],
    "4_Uzun_Vadeli_Yukumlulukler": [None, None, None, None, None, None],
    "40_Finansal_Borclar": [None, None, None, None, None, None],
    "48_Ertelenmis_Gelirler": [None, None, None, None, None, None],
    "47_Uzun_vadeli_Karsiliklar": [None, None, None, None, None, None],
    "481_Ertelenmis_Vergi_Yukumlulugu": [None, None, None, None, None, None],
    "5_Ozkaynaklar": [None, None, None, None, None, None],
    "52_Ana_Ortakliga_Ait_Ozkaynaklar": [None, None, None, None, None, None],
    "50_Odenmis_Sermaye": [None, None, None, None, None, None],
    "5_x_Yabanci_Para_Cevrim_Farklari": [None, None, None, None, None, None],
    "54_Kardan_Ayrilan_Kisitlanmis_Yedekler": [None, None, None, None, None, None],
    "57_Gecmis_Yillar_Kar_Zararlari": [None, None, None, None, None, None],
    "59_Donem_Net_Kar_Zarari": [None, None, None, None, None, None],
    "529_Diger_Ozkaynak_Kalemleri": [None, None, None, None, None, None],
    "TKAY_TOPLAM_KAYNAKLAR": [None, None, None, None, None, None],
    "241_Hedge_Dahil_Net_Yabanci_Para_Pozisyonu": [None, None, None, None, None, None],
    "11_Finansal_Yatirimlar": [None, None, None, None, None, None],
    "12_1_Finans_Sektoru_Faaliyetlerinden_Alacaklar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "22_Ticari_Alacaklar": [None, None, None, None, None, None],
    "22_1_Finans_Sektoru_Faaliyetlerinden_Alacaklar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "29_Diger_Duran_Varliklar": [None, None, None, None, None, None],
    "30_1_Diger_Finansal_Yukumlulukler": [None, None, None, None, None, None],
    "32_1_Finans_Sektoru_Faaliyetlerinden_Borclar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "33_1_Devlet_Tesvik_ve_Yardimlari": [None, None, None, None, None, None],
    "42_Ticari_Borclar": [None, None, None, None, None, None],
    "40_1_Finans_Sektoru_Faaliyetlerinden_Borclar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "43_Diger_Borclar": [None, None, None, None, None, None],
    "49_Diger_Uzun_Vadeli_Yukumlulukler": [None, None, None, None, None, None],
    "53_Azinlik_Paylari": [None, None, None, None, None, None],
    "19_1_Satis_Amaciyla_Elde_Tutulan_Duran_Varliklar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "25_2_Ozkaynak_Yontemiyle_Degerlenen_Yatirimlar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "390_1_Satis_Amacli_Siniflandirilan_Varlik_Gruplarina_Iliskin_Yukumlulukler": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "49_1_Devlet_Tesvik_ve_Yardimlari": [None, None, None, None, None, None],
    "60_Satis_Gelirleri": [None, None, None, None, None, None],
    "600_Yurt_Ici_Satislar": [None, None, None, None, None, None],
    "601_Yurt_Disi_Satislar": [None, None, None, None, None, None],
    "62_Satislarin_Maliyeti__e_": [None, None, None, None, None, None],
    "602_Ticari_Faaliyetlerden_Brut_Kar__Zarar_": [None, None, None, None, None, None],
    "BK_BRUT_KAR__ZARAR_": [None, None, None, None, None, None],
    "63_Genel_Yonetim_Giderleri__e_": [None, None, None, None, None, None],
    "631_Pazarlama__Satis_ve_Dagitim_Giderleri__e_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "630_Arastirma_ve_Gelistirme_Giderleri__e_": [None, None, None, None, None, None],
    "64_Diger_Faaliyet_Gelirleri": [None, None, None, None, None, None],
    "65_Diger_Faaliyet_Giderleri__e_": [None, None, None, None, None, None],
    "FKZ_FAALIYET_KARI__ZARARI_": [None, None, None, None, None, None],
    "64_1_Yatirim_Faaliyetlerinden_Gelirler": [None, None, None, None, None, None],
    "65_1_Yatirim_Faaliyetlerinden_Giderler__e_": [None, None, None, None, None, None],
    "FGOFKZ_FINANSMAN_GELIRI__GIDERI__ONCESI_FAALIYET_KARI__ZARARI_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "67__Esas_Faaliyet_Disi__Finansal_Gelirler": [None, None, None, None, None, None],
    "68__Esas_Faaliyet_Disi__Finansal_Giderler__e_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "SFVOKZ_SURDURULEN_FAALIYETLER_VERGI_ONCESI_KARI__ZARARI_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "SFVG_Surdurulen_Faaliyetler_Vergi_Geliri__Gideri_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "691_Donem_Vergi_Geliri__Gideri_": [None, None, None, None, None, None],
    "67_1_Ertelenmis_Vergi_Geliri__Gideri_": [None, None, None, None, None, None],
    "SFDKZ_SURDURULEN_FAALIYETLER_DONEM_KARI_ZARARI": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "DFDKZ_DURDURULAN_FAALIYETLER_DONEM_KARI_ZARARI": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "69_DONEM_KARI__ZARARI_": [None, None, None, None, None, None],
    "69_2_Azinlik_Paylari": [None, None, None, None, None, None],
    "69_3_Ana_Ortaklik_Paylari": [None, None, None, None, None, None],
    "796_Amortisman": [None, None, None, None, None, None],
    "67_2_Faiz__Ucret__Prim__Komisyon_ve_Diger_Gelirler": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "68_1_Faiz__Ucret__Prim__Komisyon_ve_Diger_Giderler__e_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "67_3_Finans_Sektoru_Faaliyetlerinden_Brut_Kar__Zarar_": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    "69_1_Ozkaynak_Yontemiyle_Deger_Yatirimlarin_Kar_Zarar__Paylar": [
        None,
        None,
        None,
        None,
        None,
        None,
    ],
}

all_fields_copy = all_fields.copy()


def fill_fields_from_excel(excel_file_path):
    # Define the structure of all_fields_copy inside the function for clarity

    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet_balance = workbook["Bilanço"]

    # Helper function to find the context for duplicate names
    def find_context(row_index):
        for i in range(row_index, 0, -1):  # Iterate backwards to find the context
            cell_value = sheet_balance[f"A{i}"].value
            if cell_value in [
                "Dönen Varlıklar",
                "Duran Varlıklar",
                "Kısa Vadeli Yükümlülükler",
                "Uzun Vadeli Yükümlülükler",
                "Özkaynaklar",
            ]:
                return cell_value
        return None

    # Iterate through each row in column A
    for row_index, row in enumerate(
        sheet_balance.iter_rows(min_row=1, max_col=1, values_only=True), start=1
    ):
        cell_value = row[0]  # Get the value of the cell in column A
        if cell_value:
            for key, value in balance_dict.items():
                if cell_value == key[0]:
                    # If there's a match, find the context for duplicates, if necessary
                    context = find_context(row_index)
                    if len(key) > 1 and context != key[1]:
                        continue  # Skip if the context doesn't match for duplicates

                    # Retrieve next 5 values in the row (columns B to F)
                    values = [
                        sheet_balance.cell(row=row_index, column=col).value
                        for col in range(2, 8)
                    ]
                    all_fields_copy[value] = (
                        values  # Store these values in all_fields_copy
                    )
                    break

    # Manually fill the 'Donem' field
    all_fields_copy["Donem"] = (
        sheet_balance["B1"].value if sheet_balance["B1"].value else [None]
    )  # Ensure this is handled correctly

    # Load the workbook and select the specified sheet
    sheet_income = workbook["Gelir Tablosu (Çeyreklik)"]

    # Iterate through each row in column A
    for row_index, row in enumerate(
        sheet_income.iter_rows(min_row=1, max_col=1, values_only=True), start=1
    ):
        cell_value = row[0]  # Get the value of the cell in column A
        if cell_value:
            for key, value in income_dict.items():
                if cell_value == key:
                    # Retrieve next 5 values in the row (columns B to F)
                    values = [
                        sheet_income.cell(row=row_index, column=col).value
                        for col in range(2, 8)
                    ]
                    all_fields_copy[value] = (
                        values  # Store these values in all_fields_copy
                    )
                    break

    # Ensure all fields are filled, even if no data was found
    for key in all_fields_copy:
        if all_fields_copy[key] is None:
            all_fields_copy[key] = [
                None
            ] * 6  # Assign a list of None if no data was present

    return all_fields_copy


calculated_ratios = {
    "B1_Donem_Kari": [None, None, None, None, None],
    "B2_Donem_Satislari": [None, None, None, None, None],
    "B3_Toplam_Borc_Buyuklugu": [None, None, None, None, None],
    "B4_Toplam_Alacak_Buyuklugu": [None, None, None, None, None],
    "B5_Aktif_Buyukluk": [None, None, None, None, None],
    "B6_Ozsermaye_Buyukluk": [None, None, None, None, None],
    "B7_Isletme_Sermaye_Buyukluk": [None, None, None, None, None],
    "B8_Odenmis_Sermaye_Buyukluk": [None, None, None, None, None],
    "L1_Cari_Oran": [None, None, None, None, None],
    "L2_Asit_Oran": [None, None, None, None, None],
    "L3_Nakit_Oran": [None, None, None, None, None],
    "L4_Hazir_Degerler_Oran": [None, None, None, None, None],
    "L5_Favok": [None, None, None, None, None],
    "L6_Faiz_Karsilama_Oran": [None, None, None, None, None],
    "L7_Stok_Bagimlilik_Oran": [None, None, None, None, None],
    "F1_Finansal_Kaldirac_Oran": [None, None, None, None, None],
    "F2_Ozkaynak_Oran": [None, None, None, None, None],
    "F3_Finansman_Oran": [None, None, None, None, None],
    "F4_Sermaye_Carpani": [None, None, None, None, None],
    "F5_Oto_Finansman": [None, None, None, None, None],
    "F6_Duran_Varlik_Sermaye_Oran": [None, None, None, None, None],
    "V1_Stok_Devir_Hizi": [None, None, None, None, None],
    "V2_Stokta_Kalma_Ortalamasi": [None, None, None, None, None],
    "V3_Alacak_Devir_Hizi": [None, None, None, None, None],
    "V4_Tahsilat_Hizi": [None, None, None, None, None],
    "V5_Borc_Devir_Hizi": [None, None, None, None, None],
    "V6_Borc_Odeme_Suresi": [None, None, None, None, None],
    "V7_Etkinlik_Hizi": [None, None, None, None, None],
    "V8_Aktif_Devir_Hizi": [None, None, None, None, None],
    "V9_Duran_Varlik_Devir_Hizi": [None, None, None, None, None],
    "V10_Ozkaynak_Devir_Hizi": [None, None, None, None, None],
    "V11_Calisma_Sermayesi_Devir_Hizi": [None, None, None, None, None],
    "V12_Net_Calisma_Sermayesi_Devir_Hizi": [None, None, None, None, None],
    "V13_Hazir_Deger_Devir_Hizi": [None, None, None, None, None],
    "K1_Ozkaynak_Karlilik": [None, None, None, None, None],
    "K2_Sermaye_Karlilik": [None, None, None, None, None],
    "K3_Aktif_Karlilik": [None, None, None, None, None],
    "K4_Ekonomik_Karlilik": [None, None, None, None, None],
    "K5_Calisma_Sermayesi_Karlilik": [None, None, None, None, None],
}


def calculate_financial_ratios(financial_data):
    # Initialize the output dictionary to store calculated ratios

    financial_data_for_ratios = financial_data.copy()
    # Helper function to calculate average or shifted values for lists
    def calculate_shifted_averages(values):
        averages = []
        for i in range(1, len(values)):
            averages.append((values[i] + values[i - 1]) / 2)
        return averages

    # Pre-calculate necessary intermediate values
    ortalama_stoklar = calculate_shifted_averages(financial_data_for_ratios["15_Stoklar"])
    donem_basi_stok = financial_data_for_ratios["15_Stoklar"][
        1:
    ]  # Exclude the last element for shift(1)
    donem_basi_borc = [
        financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i + 1]
        + financial_data_for_ratios["4_Uzun_Vadeli_Yukumlulukler"][i + 1]
        for i in range(0, 5)
    ]

    # Ensure all None values in the necessary financial fields are replaced with 0
    for key in ["12_Ticari_Alacaklar", "13_Diger_Alacaklar", "23_Diger_Alacaklar","15_Stoklar", "18_Pesin_Odenmis_Giderler","60_Satis_Gelirleri","62_Satislarin_Maliyeti__e_","63_Genel_Yonetim_Giderleri__e_","796_Amortisman","10_Nakit_ve_Nakit_Benzerleri"," "]:
        if key in financial_data_for_ratios:
            financial_data_for_ratios[key] = [
                0 if value is None else value for value in financial_data_for_ratios[key]
            ]

    for i in range(5):  # Calculating ratios for 5 periods
        # Calculate each ratio based on provided logic
        calculated_ratios["B1_Donem_Kari"][i] = financial_data_for_ratios[
            "69_DONEM_KARI__ZARARI_"
        ][i]

        calculated_ratios["B2_Donem_Satislari"][i] = financial_data_for_ratios[
            "60_Satis_Gelirleri"
        ][i]

        calculated_ratios["B3_Toplam_Borc_Buyuklugu"][i] = (
            financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
            + financial_data_for_ratios["4_Uzun_Vadeli_Yukumlulukler"][i]
        )

        calculated_ratios["B4_Toplam_Alacak_Buyuklugu"][i] = (
            financial_data_for_ratios["12_Ticari_Alacaklar"][i]
            + financial_data_for_ratios["13_Diger_Alacaklar"][i]
            + financial_data_for_ratios["23_Diger_Alacaklar"][i]
        )

        calculated_ratios["B5_Aktif_Buyukluk"][i] = (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
        )

        calculated_ratios["B6_Ozsermaye_Buyukluk"][i] = financial_data_for_ratios["5_Ozkaynaklar"][
            i
        ]

        calculated_ratios["B7_Isletme_Sermaye_Buyukluk"][i] = (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
            - financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
        )

        calculated_ratios["B8_Odenmis_Sermaye_Buyukluk"][i] = (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
            - financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
            - financial_data_for_ratios["4_Uzun_Vadeli_Yukumlulukler"][i]
        )

        calculated_ratios["L1_Cari_Oran"][i] = (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
        ) / (financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i])
        
        calculated_ratios["L2_Asit_Oran"][i] = (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            - financial_data_for_ratios["15_Stoklar"][i]
            - financial_data_for_ratios["18_Pesin_Odenmis_Giderler"][i]
        ) / (financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i])

        calculated_ratios["L3_Nakit_Oran"][i] = (
            financial_data_for_ratios["10_Nakit_ve_Nakit_Benzerleri"][i]
        ) / (financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i])

        calculated_ratios["L5_Favok"][i] = (financial_data_for_ratios["60_Satis_Gelirleri"][i]) + (financial_data_for_ratios["62_Satislarin_Maliyeti__e_"][i]) + (financial_data_for_ratios["63_Genel_Yonetim_Giderleri__e_"][i]) + (financial_data_for_ratios["796_Amortisman"][i])

        calculated_ratios["F1_Finansal_Kaldirac_Oran"][i] = (
            financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
            + financial_data_for_ratios["4_Uzun_Vadeli_Yukumlulukler"][i]
        ) / (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]

        )
        calculated_ratios["F2_Ozkaynak_Oran"][i] = (
            financial_data_for_ratios["5_Ozkaynaklar"][i]
        ) / (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
        )

        calculated_ratios["F3_Finansman_Oran"][i] = (
            financial_data_for_ratios["5_Ozkaynaklar"][i]
        ) / (
            financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
            + financial_data_for_ratios["4_Uzun_Vadeli_Yukumlulukler"][i]
        )
        
        calculated_ratios["F4_Sermaye_Carpani"][i] = (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
        ) / (financial_data_for_ratios["5_Ozkaynaklar"][i])

        calculated_ratios["V1_Stok_Devir_Hizi"][i] = (
            -financial_data_for_ratios["62_Satislarin_Maliyeti__e_"][i]
        ) / (ortalama_stoklar[i])

        calculated_ratios["V3_Alacak_Devir_Hizi"][i] = (
            financial_data_for_ratios["60_Satis_Gelirleri"][i]
        ) / financial_data_for_ratios["12_Ticari_Alacaklar"][i]

        calculated_ratios["V8_Aktif_Devir_Hizi"][i] = (
            financial_data_for_ratios["60_Satis_Gelirleri"][i]
        ) / (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
        )

        calculated_ratios["V11_Calisma_Sermayesi_Devir_Hizi"][i] = (
            financial_data_for_ratios["60_Satis_Gelirleri"][i]
        ) / (
            financial_data["1_Donen_Varliklar"][i]
            - financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
        )

        calculated_ratios["K1_Ozkaynak_Karlilik"][i] = (
            financial_data_for_ratios["69_DONEM_KARI__ZARARI_"][i]
        ) / financial_data_for_ratios["5_Ozkaynaklar"][i]

        calculated_ratios["K2_Sermaye_Karlilik"][i] = (financial_data_for_ratios["69_DONEM_KARI__ZARARI_"][i]) / (financial_data_for_ratios["1_Donen_Varliklar"][i] + financial_data_for_ratios["2_Duran_Varliklar"][i] - financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i] - financial_data_for_ratios["4_Uzun_Vadeli_Yukumlulukler"][i])

        calculated_ratios["K3_Aktif_Karlilik"][i] = (
            financial_data_for_ratios["69_DONEM_KARI__ZARARI_"][i]
        ) / (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            + financial_data_for_ratios["2_Duran_Varliklar"][i]
        )
        
        calculated_ratios["K5_Calisma_Sermayesi_Karlilik"][i] = (
            financial_data_for_ratios["69_DONEM_KARI__ZARARI_"][i]
        ) / (
            financial_data_for_ratios["1_Donen_Varliklar"][i]
            - financial_data_for_ratios["3_Kisa_Vadeli_Yukumlulukler"][i]
        )

    return calculated_ratios


def assign_score(ratios, sector):
    """
    Assign a score to a firm based on financial ratios and sector.

    Args:
        ratios (dict): A dictionary containing financial ratios. The keys should match the predefined ratios.
        sector (str): The sector of the firm.
        ratio_type (str): The type of ratios to use ('liquidity', 'size', 'financial_structure', 'efficiency', 'profitability').

    Returns:
        int: The assigned score out of 10.
    """
    # Check if the sector name exists in the predefined sectors
    sector = sector.lower()
    allowed_sectors = [
        "kimya",
        "imalat",
        "toptan",
        "gida",
        "insaat",
        "yatirim",
        "gayrimenkul",
        "enerji",
        "hizmet",
        "bilisim",
        "anametal",
        "tekstil",
    ]
    if sector not in allowed_sectors:
        raise ValueError(
            "Invalid sector. Allowed sectors: kimya, imalat, toptan, gida, insaat, yatirim, gayrimenkul, enerji, hizmet, bilisim, anametal, tekstil"
        )

    # Check if the ratio type exists in the predefined list
    allowed_types = [
        "buyukluk",
        "likidite",
        "finansal_yapi",
        "varlik_yonetim",
        "karlilik",
    ]

    Scores = {}
    match_dict = {
        "buyukluk": ["Buyukluk_Skor", "B"],
        "likidite": ["Likidite_Skor", "L"],
        "finansal_yapi": ["Finansal_Yapi_Skor", "F"],
        "varlik_yonetim": ["Varlik_Kullanim_Skor", "V"],
        "karlilik": ["Karlilik", "K"],
    }
    for i in allowed_types:
        # Define the paths for ideal values and quantile boundaries dictionaries based on the ratio_type
        ideal_values_path = f"data/dicts/ideal_values_dict_{i}.json"
        quantile_boundaries_path = f"data/dicts/quantile_boundaries_dict_{i}.json"

        # Load the ideal values and quantile boundaries dictionaries
        with open(ideal_values_path, "r") as iv_file, open(
            quantile_boundaries_path, "r"
        ) as qb_file:
            ideal_values_dict = json.load(iv_file)
            quantile_boundaries_dict = json.load(qb_file)

        # Check if the sector is in the dictionaries
        if (
            f"ideal_values_{sector}" not in ideal_values_dict
            or f"quantile_boundaries_{sector}" not in quantile_boundaries_dict
        ):
            raise ValueError(f"No data found for sector: {sector}")

        # Get the ideal values and quantile boundaries for the sector
        ideal_values = ideal_values_dict[f"ideal_values_{sector}"]
        quantile_boundaries = quantile_boundaries_dict[f"quantile_boundaries_{sector}"]

        specified_ratios = {}

        for key, value in ratios.items():
            if key.startswith(match_dict[i][1]):
                specified_ratios[key] = value

        specified_ratios = {
            key: value for key, value in specified_ratios.items() if None not in value
        }

        specified_ratios = {key: value[0] for key, value in specified_ratios.items()}
        # Ensure the number of ratios and ideal values match
        if len(specified_ratios) != len(ideal_values):
            raise ValueError(
                f"The number of ratios provided ({len(specified_ratios)}) does not match the expected number of ratios ({len(ideal_values)})"
            )

        # Calculate the distance from the reference point
        distance = np.sqrt(
            sum((np.array(list(specified_ratios.values())) - ideal_values) ** 2)
        )

        n_clusters = 10

        # Determine the cluster based on quantile boundaries
        for a, boundary in enumerate(quantile_boundaries):
            if distance <= boundary:
                result = 10 - a  # Subtract 1 because quantiles start at 0
                Scores[match_dict[i][0]] = result
                break
            elif distance > max(quantile_boundaries):
                result = 10 - (
                    n_clusters - 1
                )  # Assign to the last cluster if distance exceeds all boundaries
                Scores[match_dict[i][0]] = result
                break
            else:
                continue

    return Scores

def count_rows_in_each_sheet(excel_file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)
    
    # Dictionary to hold the sheet names and their respective row counts
    sheet_row_counts = {}

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Get the number of rows in the sheet
        row_count = sheet.max_row - 1
        # Add the sheet name and row count to the dictionary
        sheet_row_counts[sheet_name] = row_count

    return sheet_row_counts


@app.post("/analyze/")
async def analyze_financials(
    excel_file: UploadFile = File(...),
    sector_name: str = None,
    token: str = Depends(valid_api_token),
):

    if (
        excel_file.content_type
        != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        raise HTTPException(
            status_code=400, detail="Invalid file format. Please upload an Excel file."
        )

    try:

        # Create a temporary file to save the uploaded file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            # Write the uploaded file's contents to the temporary file
            content = await excel_file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name

        # Here, you would call your modified fill_fields_from_excel() function
        financial_data = fill_fields_from_excel(temp_file_path)

        # Then, you calculate financial ratios
        ratios = calculate_financial_ratios(financial_data)

        # Finally, you assign a score
        scores = assign_score(ratios, sector_name)

        summary = count_rows_in_each_sheet(temp_file_path)

        # Prepare the final output
        result = {
            "Donem": financial_data["Donem"],
            "Finansal_THP_Alanlari": financial_data,
            "Rasyolar": ratios,
            "Skorlar": scores,
            "Summary": summary,
        }

        return JSONResponse(content=result)

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
